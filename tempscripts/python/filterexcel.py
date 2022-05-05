#!/usr/bin/python
#coding=utf-8
# 使用openpyxl来过滤excel中的数据
# 如果没有openpyxl，请使用pip install openpyxl安装

import time
import sys,getopt
from openpyxl import load_workbook
from openpyxl import Workbook

def usage():
  helpmsg = """
使用方法:python {} -x excel文件 -f 筛选文件 -t 筛选标题行号 -k 筛选关键字 -r 是否反向筛选
         -x --xlsx excel文件路径
         -f --filter 筛选文件路径
         -t --title excel文件的筛选行号
         -k --keyword 筛选关键字
         可选-r --reverse 是否反选
""".format(sys.argv[0])
  print(helpmsg)
  sys.exit(0)

def checkargv(argv):
  xlsx = ""
  filter = ""
  title = ""
  keyword = ""
  reverse_chosen = False

  if len(sys.argv) < 4:
    usage()

  try:
    opts, args = getopt.getopt(argv, "hx:f:t:k:r", ["help", "xlsx=", "filter=", "title=", "keyword="])
  except getopt.GetoptError:
    print("输入错了参数")
    usage()

  for opt, arg in opts:
    if opt in ("-h", "--help"):
      usage()
    elif opt in ("-x", "--xlsx"):
      xlsx = arg
    elif opt in ("-f", "--filter"):
      filter = arg
    elif opt in ("-t", "--title"):
      title = arg
    elif opt in ("-k", "--keyword"):
      keyword = arg
    elif opt in ("-r", "--reverse"):
      reverse_chosen = True

  return xlsx, filter, title, keyword, reverse_chosen

class ExcelOp(object):
  # 初始化载入工作簿和sheet
  def __init__(self, file):
    self.file = file
    self.wb = load_workbook(self.file)
    self.sheet = self.wb.get_sheet_names()
    self.ws = self.wb[self.sheet[0]]

  # 获取表格的总行数和总列数
  def get_row_col_num(self, type):
    if type == 'row':
      rows = self.ws.max_row
      return rows
    elif type == 'col':
      columns = self.ws.max_column
      return columns

  # 获取某行的所有值
  def get_row_value(self, row):
    columns = self.ws.max_column
    row_data = []
    for i in range(1, columns + 1):
      cell_value = self.ws.cell(row=row, column=i).value
      row_data.append(cell_value)
    return row_data

# 屏幕信息输出带时间戳
def printmeg(megstr):
  nowt = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
  print("{} {}".format(nowt, megstr))

# 筛选出列表信息的导出方法
def tranfilter(rawlist):

  filet = time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime())
  dest_file = "./excelfilterresult-{}.xlsx".format(filet)
  wb = Workbook()
  ws1 = wb.active
  ws1.title = "filterresult"

  # 写数据，筛选数据
  rownum = 1
  for row in rawlist:
    for col in range(1, len(row) + 1):
      _ = ws1.cell(column = col, row = rownum, value = "{0}".format(row[col - 1]))
    rownum = rownum + 1

  wb.save(dest_file)
  megstr = "完成，结果请查看{}".format(dest_file)
  printmeg(megstr)

if __name__ == "__main__":

  if len(sys.argv) < 2:
    usage()
  else:
    xlsx, filter, title, keyword, reverse_chosen = checkargv(sys.argv[1:])

    try:
      sourcews = ExcelOp(xlsx)
      sourcerow = sourcews.get_row_col_num('row')
      sourcecol = sourcews.get_row_col_num('col')

      if sourcerow == 0:
        megstr = "错误，{}的sheet1是空表".format(xlsx)
        printmeg(megstr)
        exit(1)
      
      # 读取筛选文件
      f = open(filter, 'r')
      filtercontlist = f.read().split('\n')
      # 初始化筛选数据
      filterdatalist = []
      # 获取标题数据列表
      titlelist = sourcews.get_row_value(int(title))
      # 获取筛选关键字所在的列号
      filtercol = titlelist.index(keyword)
      # 循环添加筛选信息
      for row in range(1, sourcerow + 1):
        rowlist = sourcews.get_row_value(row)

        if reverse_chosen:
          if str(rowlist[filtercol]) not in filtercontlist:
            filterdatalist.append(rowlist)
        else:
          if str(rowlist[filtercol]) in filtercontlist:
            filterdatalist.append(rowlist)

      
      # 输出筛选后的结果
      if filterdatalist == []:
        megstr = "错误，没找到对应的数据，请检查列号"
        printmeg(megstr)
      else:
        # 这步是为了插入标题
        if titlelist not in filterdatalist:
          filterdatalist.insert(0, titlelist)
        # 把筛选结果写入新文件
        tranfilter(filterdatalist)
    except:
      raise Exception
